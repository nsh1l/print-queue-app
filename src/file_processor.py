"""
File type detection and processing for XLSX, XLS, PDF.
"""
import fitz  # PyMuPDF
import xlrd
import openpyxl
import subprocess
import tempfile
import os
import uuid
from pathlib import Path
from typing import Literal


FileType = Literal["xlsx", "xls", "pdf"]


def detect_file_type(path: Path) -> FileType | None:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return "xlsx"
    if suffix == ".xls":
        return "xls"
    if suffix == ".pdf":
        return "pdf"
    return None


def get_sheet_names(path: Path, file_type: FileType) -> list[str]:
    """Return sheet names for preview."""
    if file_type == "xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    if file_type == "xls":
        wb = xlrd.open_workbook(path)
        names = wb.sheet_names()
        wb.release_resources()
        return names
    if file_type == "pdf":
        doc = fitz.open(path)
        count = doc.page_count
        doc.close()
        return [f"Page {i+1}" for i in range(count)]
    return []


def render_pdf_page(path: Path, page: int, dpi: int = 150) -> bytes:
    """Render a PDF page as PNG bytes."""
    doc = fitz.open(path)
    page_obj = doc[page]
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page_obj.get_pixmap(matrix=mat)
    data = pix.tobytes("png")
    doc.close()
    return data


def xlsx_to_csv_preview(path: Path, sheet_index: int = 0, max_rows: int = 20) -> str:
    """Extract a sheet as CSV preview string."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.worksheets[sheet_index]
    lines = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= max_rows:
            lines.append("... (truncated)")
            break
        lines.append(",".join(str(c) if c is not None else "" for c in row))
    wb.close()
    return "\n".join(lines)


def xls_to_csv_preview(path: Path, sheet_index: int = 0, max_rows: int = 20) -> str:
    """Extract a sheet as CSV preview string."""
    wb = xlrd.open_workbook(path)
    sheet = wb.sheet_by_index(sheet_index)
    lines = []
    for i in range(min(max_rows, sheet.nrows)):
        row = sheet.row_values(i)
        lines.append(",".join(str(c) for c in row))
    if sheet.nrows > max_rows:
        lines.append("... (truncated)")
    wb.release_resources()
    return "\n".join(lines)


def pdf_page_count(path: Path) -> int:
    doc = fitz.open(path)
    count = doc.page_count
    doc.close()
    return count
